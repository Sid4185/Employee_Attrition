#!/usr/bin/env python
# coding: utf-8

# # importing JSON and reading JSON data from a file

# In[1]:


import json
with open('data.json','r') as file:
    data= json.load(file)
print(data)


# # writing JSON data to a file

# In[ ]:


open("output.json",'w') as file:
    json.dump(data, file)


json_string= json.dumps(data, ensure_ascii=False)   #ascii sorting
json_string= json.dumps(data, sort_keys=True, indent=4)  #indentation


# # creating the API link using API key and other credentials 

# In[ ]:


pip install requests pandas


# In[ ]:


import requests
api_url = "desired url"    #put the real desired url of the API

# authentication for the API
headers = {"Authorization" : "Bearer api_key"}    #put the real API key here
response = requests.get(api_url, headers == headers)

response = requests.post(url, json = data)

#checking if the request was succesful or not
if response.status_code==200:
    data = response.json()
else:
    print("Error:", response.status_code)
    data=[]


# # Handling API rate limits and errors

# In[ ]:


try:
    response= requests.get(api_url, headers = headers)
response.raise_for_status()
except requests.exceptions.HTTPError as err:
    print(f"HTTP Error occured: {err}")
except exception as err:
    print(f"Other error occured: {err}")


# # converting API data into a dataframe to be used

# In[ ]:


import pandas as pd
df=pd.DataFrame(data)  #since JSON saves data in form of a dictionary, we can directly convert JSON data into a dataframe.
df.head()


# In[ ]:


#if we want data in CSV format

url= "https://required_url/data.csv"
df=pd.read_csv(url)


# In[ ]:





# # Importing data and preprocessing

# In[ ]:


import pandas as pd
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
import seaborn as sns


from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


# In[ ]:


df=pd.read_csv("HR-Employee-Data.csv")
df.head()


# In[ ]:


df.columns


# In[ ]:


df.isnull().sum()


# In[ ]:


df.dtypes


# In[ ]:


columns_to_remove = ['EmployeeCount', 'EmployeeNumber', 'Over18', 'StandardHours']

# Remove the columns
df = df.drop(columns=columns_to_remove)
df.head()


# In[ ]:


df['Attrition'] = df['Attrition'].apply(lambda x: 1 if x == 'Yes' else 0)


# In[ ]:


# List of categorical columns excluding 'Attrition'
categorical_cols = ['BusinessTravel', 'Department', 'EducationField', 'Gender', 'JobRole', 'MaritalStatus', 'OverTime']

# Perform one-hot encoding
df_encoded = pd.get_dummies(df, columns=categorical_cols, drop_first=True)


# In[ ]:


# List of numerical columns to be standardized
numerical_cols = df_encoded.select_dtypes(include=['int64', 'float64']).columns.tolist()
numerical_cols.remove('Attrition')  # Exclude the target variable


# In[ ]:


scaler = StandardScaler()

# Standardize the numerical features
df_encoded[numerical_cols] = scaler.fit_transform(df_encoded[numerical_cols])



# In[ ]:


# Display the first few rows of the standardized dataframe
df_encoded.head()


# In[ ]:


df_encoded.hist(bins=30, figsize=(20, 15))
plt.suptitle('Distribution of Numerical Features')
plt.xlabel('Feature Values')  # add x-axis label
plt.ylabel('Frequency')  # add y-axis label
plt.show()


# In[ ]:


# Plot the correlation matrix
plt.figure(figsize=(25, 20))
sns.heatmap(df_encoded.corr(), annot=True, fmt=".2f")
plt.title('Correlation Matrix')
plt.show()


# In[ ]:


# Plot the distribution of the target variable 'Attrition'
plt.figure(figsize=(6, 4))
sns.countplot(x='Attrition', data=df)
plt.title('Distribution of Attrition')
plt.show()


# In[ ]:


# Plot relationships between some selected features and 'Attrition'
selected_features = ['Age', 'MonthlyIncome', 'TotalWorkingYears', 'JobSatisfaction']
for feature in selected_features:
    plt.figure(figsize=(6, 4))
    sns.boxplot(x='Attrition', y=feature, data=df)
    plt.title(f'{feature} vs Attrition')
    plt.show()


# In[ ]:


from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix

# Separate features (X) and target variable (y)
X = df_encoded.drop('Attrition', axis=1)
y = df_encoded['Attrition']

# Split the data into training and testing sets (80% train, 20% test)
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Initialize the Logistic Regression model
model = LogisticRegression(max_iter=1000, random_state=42)

# Train the model
model.fit(X_train, y_train)

# Predict on the test set
y_pred = model.predict(X_test)

# Evaluate the model
accuracy = accuracy_score(y_test, y_pred)
print(f'Accuracy: {accuracy:.2f}\n')

print('Classification Report:')
print(classification_report(y_test, y_pred))

print('Confusion Matrix:')
print(confusion_matrix(y_test, y_pred))


# In[ ]:


from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix

# Initialize the Random Forest model
rf_model = RandomForestClassifier(random_state=42)

# Train the model
rf_model.fit(X_train, y_train)

# Predict on the test set
y_pred_rf = rf_model.predict(X_test)

# Evaluate the model
accuracy_rf = accuracy_score(y_test, y_pred_rf)
print(f'Random Forest Accuracy: {accuracy_rf:.2f}\n')

print('Random Forest Classification Report:')
print(classification_report(y_test, y_pred_rf))

print('Random Forest Confusion Matrix:')
print(confusion_matrix(y_test, y_pred_rf))


# In[ ]:


# Extract feature importances
feature_importances = rf_model.feature_importances_

# Create a DataFrame to display feature importances
feature_importances_df = pd.DataFrame({'Feature': X.columns, 'Importance': feature_importances})
feature_importances_df = feature_importances_df.sort_values(by='Importance', ascending=False)

# Display the top 10 features by importance
print('Top 10 Features by Importance:')
print(feature_importances_df.head(10))

feat=feature_importances_df


# In[ ]:


from sklearn.model_selection import GridSearchCV

# Define the parameter grid
param_grid = {
    'n_estimators': [100, 200, 300],
    'max_depth': [None, 10, 20],
    'min_samples_split': [2, 5, 10],
    'min_samples_leaf': [1, 2, 4]
}

# Initialize GridSearchCV
grid_search = GridSearchCV(estimator=RandomForestClassifier(random_state=42),
                           param_grid=param_grid,
                           scoring='accuracy',
                           cv=5,
                           verbose=1,
                           n_jobs=-1)

# Fit GridSearchCV
grid_search.fit(X_train, y_train)

# Print the best parameters and best score
print('Best Parameters:', grid_search.best_params_)
print('Best CV Score:', grid_search.best_score_)


# In[ ]:


from sklearn.metrics import accuracy_score, classification_report, confusion_matrix

# Initialize the tuned Random Forest model with best parameters
best_rf_model = RandomForestClassifier(n_estimators=200,
                                       max_depth=None,
                                       min_samples_leaf=1,
                                       min_samples_split=2,
                                       random_state=42)

# Train the model on the entire training data (X_train, y_train)
best_rf_model.fit(X_train, y_train)

# Predict on the test set
y_pred_rf_tuned = best_rf_model.predict(X_test)

# Evaluate the model
accuracy_rf_tuned = accuracy_score(y_test, y_pred_rf_tuned)
print(f'Tuned Random Forest Accuracy: {accuracy_rf_tuned:.2f}\n')

print('Tuned Random Forest Classification Report:')
print(classification_report(y_test, y_pred_rf_tuned))

print('Tuned Random Forest Confusion Matrix:')
print(confusion_matrix(y_test, y_pred_rf_tuned))



# In[ ]:


import xgboost as xgb
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix

# Initialize the XGBoost model
xgb_model = xgb.XGBClassifier(objective='binary:logistic', random_state=42)

# Train the model
xgb_model.fit(X_train, y_train)

# Predict on the test set
y_pred_xgb = xgb_model.predict(X_test)

# Evaluate the model
accuracy_xgb = accuracy_score(y_test, y_pred_xgb)
print(f'XGBoost Accuracy: {accuracy_xgb:.2f}\n')

print('XGBoost Classification Report:')
print(classification_report(y_test, y_pred_xgb))

print('XGBoost Confusion Matrix:')
print(confusion_matrix(y_test, y_pred_xgb))


# In[ ]:


from sklearn.ensemble import VotingClassifier

# Initialize the Voting Classifier with Random Forest and XGBoost
voting_model = VotingClassifier(estimators=[
    ('rf', RandomForestClassifier(n_estimators=200, max_depth=None, min_samples_leaf=1, min_samples_split=2, random_state=42)),
    ('xgb', xgb.XGBClassifier(objective='binary:logistic', random_state=42))
], voting='soft')  # 'soft' voting for probabilities

# Train the Voting Classifier
voting_model.fit(X_train, y_train)

# Predict on the test set
y_pred_voting = voting_model.predict(X_test)

# Evaluate the Voting Classifier
accuracy_voting = accuracy_score(y_test, y_pred_voting)
print(f'Voting Classifier Accuracy: {accuracy_voting:.2f}\n')

print('Voting Classifier Classification Report:')
print(classification_report(y_test, y_pred_voting))

print('Voting Classifier Confusion Matrix:')
print(confusion_matrix(y_test, y_pred_voting))


# In[ ]:


# Example: Random Forest feature importance
feature_importances = rf_model.feature_importances_


# In[ ]:


from sklearn.inspection import permutation_importance

# Example: Permutation Importance
perm_importance = permutation_importance(rf_model, X_test, y_test)


# In[ ]:


# Example: Visualizing feature importance from Random Forest
import matplotlib.pyplot as plt

plt.figure(figsize=(20, 14))
plt.barh(X_train.columns, rf_model.feature_importances_)
plt.xlabel('Importance')
plt.ylabel('Feature')
plt.title('Feature Importance from Random Forest')
plt.show()


# In[ ]:


# Initialize empty lists to store results
results_rf = []
results_xgb = []


# Make predictions on the test set using XGBoost
y_pred_rf = rf_model.predict(X_test)
results_rf.append(y_pred_rf.tolist())  # Store predictions in results_xgb list

# Make predictions on the test set using XGBoost
y_pred_xgb = xgb_model.predict(X_test)
results_xgb.append(y_pred_xgb.tolist())  # Store predictions in results_xgb list


# # storing result in excel file

# In[ ]:


import openpyxl
results =   [results_rf,results_xgb,]
results_df = pd.DataFrame(results)

# Setting the path to the output Excel file
Output= (r"C:\Users\dashs\OneDrive\Desktop\Output.xlsx")

# Writing the DataFrame to an Excel file
results_df.to_excel(Output, index=False)
print(results)


# # saving all the plots

# In[ ]:


# Save the plot to a file (e.g., 'histogram.png')
plt.savefig('histogram.png')

# Create a new Excel file using openpyxl
wb = Workbook()
ws = wb.active

# Add the plot to the Excel file
img = Image('histogram.png')
ws.add_image(img, 'A1')

# Add the data to the Excel file (optional)
rows = dataframe_to_rows(df_encoded, index=False, header=True)
for row in rows:
    ws.append(row)

# Save the Excel file
wb.save(Output)


# # inserting all the plots into output excel file

# In[ ]:





# In[ ]:





# In[ ]:


with pd.ExcelWriter(Output) as writer:
    feat.to_excel(writer, index=False)


# In[ ]:


pip install lifelines


# In[ ]:


pip install pyxll


# In[ ]:


import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import lifelines  # for Kaplan-Meier estimate
import pyxll as px

# Sort the dataframe by Monthly Income
df_sorted = df.sort_values(by='MonthlyIncome')


# Create a new Excel file
wb = Workbook()
ws = wb.active

# Write the sorted data to the Excel file
rows = dataframe_to_rows(df_sorted, index=False, header=True)
for row in rows:
    ws.append(row)
row_num=1

# Insert a new column to the right of the data
ws.insert_cols(ws.max_column + 1)
ws.insert_cols(ws.max_column + 1)
ws.insert_cols(ws.max_column + 1)


# Add feature importances to the Excel file
feat_row_num = row_num + 10  # add some space between the plots and the feature importances
feat_header_row = [f'Feature Importances']
ws.append(feat_header_row)
feat_rows = dataframe_to_rows(feat, index=False, header=False)
for row in feat_rows:
    ws.append(row)

# Create a histogram plot for each numerical column
numerical_cols = df_sorted.select_dtypes(include=[int, float]).columns
row_num = 1  # start from the first row

for col in numerical_cols:
    # Create a histogram plot
    fig, ax = plt.subplots(figsize=(6, 4))
    n, bins, patches = ax.hist(df_sorted[col],bins=30)
    
    '''df_sorted[col].hist(bins=30, ax=ax)   #monotone'''
    
    '''# Create a fading color effect using a colormap
    cmap = plt.get_cmap('Blues')  # choose a colormap (e.g., Blues, Reds, Greens)
    for i, patch in enumerate(patches):
        color = cmap(i / len(patches))  # fade from 0 to 1
        patch.set_facecolor(color)'''
    
    '''cmap = plt.get_cmap('rainbow')  # use a rainbow colormap
    for i, patch in enumerate(patches):
        color = cmap(i / len(patches))  # fade from 0 to 1
        patch.set_facecolor(color)'''
    
    cmap = plt.get_cmap('viridis')  # use a perceptually uniform colormap useful for the colour blind
    for i, patch in enumerate(patches):
        color = cmap(i / len(patches))  # fade from 0 to 1
        patch.set_facecolor(color)
        
    ax.set_title(f'Histogram of {col}')
    ax.set_xlabel(f'{col} Values')  # add x-axis label
    ax.set_ylabel(f'Count of {col}')  # add y-axis label
    
    mean_val = df_sorted[col].mean()
    ax.axvline(mean_val, color='black', linestyle='dashed', linewidth=2, label='Mean')  #shows mean of the data
   
    
    # Save the plot to a BytesIO object
    buf = BytesIO()
    fig.savefig(buf, format='png')
    buf.seek(0)
    
    # Add the plot to the new column
    img = Image(buf)
    col_letter = get_column_letter(ws.max_column+3)  # get the column letter
    ws.add_image(img, f'{col_letter}{row_num}')  # add the image to the new column
    
    plt.close(fig)

    
    # Move to the next row
    row_num += 30


    
# Create a Kaplan-Meier plot for TotalWorkingYears
kmf = lifelines.KaplanMeierFitter()
kmf.fit(df_sorted['TotalWorkingYears'], event_observed=df_sorted['JobSatisfaction'])

fig, ax = plt.subplots(figsize=(6, 4))
kmf.plot(ax=ax)
ax.set_title('Kaplan-Meier Curve of Total Working Years')
ax.set_xlabel('Time')
ax.set_ylabel('Survival Probability')

# Save the plot to a BytesIO object
buf = BytesIO()
fig.savefig(buf, format='png')
buf.seek(0)

# Add the plot to the new column
img = Image(buf)
col_letter = get_column_letter(ws.max_column+3)  # get the column letter
ws.add_image(img, f'{col_letter}{row_num}')  # add the image to the new column
plt.close(fig)

row_num += 30

# Create a Kaplan-Meier plot for YearsAtCompany
kmf = lifelines.KaplanMeierFitter()
kmf.fit(df_sorted['YearsAtCompany'], event_observed=df_sorted['JobSatisfaction'])

fig, ax = plt.subplots(figsize=(6, 4))
kmf.plot(ax=ax)
ax.set_title('Kaplan-Meier Curve of Years at Company')
ax.set_xlabel('Time')
ax.set_ylabel('Survival Probability')

# Save the plot to a BytesIO object
buf = BytesIO()
fig.savefig(buf, format='png')
buf.seek(0)

# Add the plot to the new column
img = Image(buf)
col_letter = get_column_letter(ws.max_column+3)  # get the column letter
ws.add_image(img, f'{col_letter}{row_num}')  # add the image to the new column
plt.close(fig)


# Add summary statistics for each numerical column
summary_stats = df_sorted.select_dtypes(include=[int, float]).describe()
summary_rows = dataframe_to_rows(summary_stats, index=True, header=True)
ws.append(['Summary Statistics'])
for row in summary_rows:
    ws.append(row)
    
# Save the Excel file
wb.save(Output)



# In[ ]:





# In[ ]:





# import bokeh.plotting as bp
# import numpy as np
# from bokeh.models import ColumnDataSource, HoverTool
# from bokeh.io import show
# from bokeh.embed import file_html
# from bokeh.plotting import figure, output_file, save
# from openpyxl import Workbook
# from openpyxl.utils.cell import get_column_letter
# from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter
# import pandas as pd
# from openpyxl.worksheet.hyperlink import Hyperlink
# 
# numerical_cols = df_sorted.select_dtypes(include=[int, float]).columns
# 
# '''wb = Workbook()
# ws = wb.active'''
# 
# 
# html_files=[]
# 
# for col in numerical_cols:
#     # Create a histogram plot
#     hist, bins = np.histogram(df_sorted[col], bins=30)
#     delays = pd.DataFrame({'arr_delay': hist, 'left': bins[:-1], 'right': bins[1:]})
#     
#     #add a link column to the dataframe
#     delays['link'] = f"[Histogram of {col}](histogram_{col}.html)"
#     
#     # Create a ColumnDataSource
#     src = ColumnDataSource(delays)
# 
#     # Create a figure (Bokeh interactive plot)
#     p = bp.figure(title=f'Histogram of {col}', x_axis_label=f'{col} Values', y_axis_label='Count')
#     p.quad(source=src, bottom=0, top='arr_delay', left='left', right='right', fill_color='red', line_color='black')
# 
#     # Add a HoverTool
#     hover = HoverTool(tooltips=[('Delay Interval Left', '@left'), ('Count', '@arr_delay')])
#     p.add_tools(hover)
# 
#     # Save the plot to an HTML file
#     html_file=f"histogram_{col}.html"
#     output_file(html_file, title=f"Histogram of {col}")
#     show(p)
#     save(p, filename= html_file)
#     
#     '''html_files.append(html_file)
#     
#     # write the link to the Excel file
#     hyperlink = Hyperlink(display=f"Histogram of {col}", target=f"file://{html_file}")
#     ws.cell(row=ws.max_row + 1, column=1).hyperlink = hyperlink
#     ws.cell(row=ws.max_row, column=1).font = Font(underline='single', color='0563C1')'''
# 
# #wb.save(Output)

# In[ ]:





# In[ ]:


import pandas as pd
from bokeh.plotting import figure, show, output_file
from bokeh.embed import components
from bokeh.models import HoverTool



# Create a new column to store the HTML links
df_sorted['URL Plots'] = ''

# Loop through each row and create a Bokeh plot
for index, row in df.iterrows():
    # Create a figure
    p = figure(title="Delay Interval Left vs Count", x_axis_label='Delay Interval Left', y_axis_label='Count')
    p.circle([row['left']], [row['arr_delay']], size=10, color="navy", alpha=0.5)

    # Add a hover tool
    hover = HoverTool(tooltips=[
        ('Delay Interval Left', '@x'),
        ('Count', '@y'),
    ])
    p.add_tools(hover)

    # Save the plot to an HTML file
    output_file("plot_{}.html".format(index), title="plot_{}".format(index))
    show(p)

    # Generate the HTML link
    html_link = '<a href="plot_{}.html" target="_blank" rel="noopener noreferrer">Plot {}</a>'.format(index, index)

    # Update the 'URL Plots' column
    df_sorted.at[index, 'URL Plots'] = html_link

# Save the updated dataframe to a new CSV file
df_sorted.to_csv('output.csv', index=False)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:


import bokeh.plotting as bp
import numpy as np
from bokeh.models import ColumnDataSource, HoverTool
from bokeh.io import show
from bokeh.embed import file_html
from bokeh.plotting import figure, output_file, save
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd

numerical_cols = df_sorted.select_dtypes(include=[int, float]).columns

# Create a new column to store the HTML links
df_sorted['URL Plots'] = ''

html_files = []

for col in numerical_cols:
    # Create a histogram plot
    hist, bins = np.histogram(df_sorted[col], bins=30)
    delays = pd.DataFrame({'arr_delay': hist, 'left': bins[:-1], 'right': bins[1:]})
    
    # Create a ColumnDataSource
    src = ColumnDataSource(delays)

    # Create a figure (Bokeh interactive plot)
    p = bp.figure(title=f'Histogram of {col}', x_axis_label=f'{col} Values', y_axis_label='Count')
    p.quad(source=src, bottom=0, top='arr_delay', left='left', right='right', fill_color='red', line_color='black')

    # Add a HoverTool
    hover = HoverTool(tooltips=[('Delay Interval Left', '@left'), ('Count', '@arr_delay')])
    p.add_tools(hover) 

    # Save the plot to an HTML file
    html_file = f"histogram_{col}.html"
    output_file(html_file, title=f"Histogram of {col}")
    show(p)
    save(p, filename=html_file)
    
    # Generate the HTML link
    html_link = f'<a href="{html_file}" target="_blank" rel="noopener noreferrer">Histogram of {col}</a>'
    
    # Update the 'URL Plots' column
    df_sorted.loc[df_sorted[col].index, 'URL Plots'] = html_link

# Save the updated dataframe to a new CSV file
df_sorted.to_csv('output.csv', index=False)


# In[ ]:





# In[ ]:





# In[ ]:




